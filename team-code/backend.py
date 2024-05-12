import os
import sys
import xlwings as xw
from openpyxl import load_workbook

class ExcelManipulator:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.workbook = load_workbook(filename=file_path)
        self.sheet_name = sheet_name
        self.sheet = self.workbook.active if sheet_name is None else self.workbook[sheet_name]

    def read_cell(self, cell):
        return self.sheet[cell].value

    def write_cell(self, cell, value):
        self.sheet[cell].value = value

    def save(self):
        self.workbook.save(filename=self.file_path)

    def add_value_to_input_column(self, input_column, value):
        for row in range(1, self.sheet.max_row + 1):
            cell = self.sheet[f"{input_column}{row}"]
            if cell.value is not None:
                continue
            cell.value = value
            return

    def get_value_from_output_cell(self, output_cell):
        return self.sheet[output_cell].value

    def vlookup(self, lookup_value, table_range, result_column):
        for row in self.sheet.iter_rows(min_row=self.sheet[table_range.split(':')[0]].row,
                max_row=self.sheet[table_range.split(':')[1]].row,
                min_col=self.sheet[table_range.split(':')[0]].column,
                max_col=self.sheet[table_range.split(':')[1]].column):
            if str(row[0].value) == lookup_value:
                return row[result_column - 1].value
        excel.save()
        return None
    
    def update_cell_value(self, cell, value):
        self.sheet[cell].value = value
        self.workbook.save(filename=self.file_path)



class SIISOTEA:
    def __init__(self, df, excel) -> None:
        self.df = df
        self.openpyxlsExcel = excel

        self.AnnualOperatingHour = 6570.00

        self.equipEntryRowIndex = self.rawMaterialRowIndex = self.productDataRowIndex = 1
        self.openpyxl_equipEntryRow = self.equipEntryRowIndex + 2
        self.openpyxl_rawMaterialRow = self.rawMaterialRowIndex + 2
        self.openpyxl_productDataRow = self.productDataRowIndex + 2

        # Equicpment Data Cost
        self.processAreaIndex = lambda: [self.equipEntryRowIndex, 12]
        self.processAreaNamesIndex = lambda: [self.equipEntryRowIndex, 13]
        self.installedCostIndex = lambda: [self.equipEntryRowIndex, 14]
        self.InsideBatteryLimitIndex = lambda: [2, 15]

        # Raw Material / Utility
        self.rawMaterialNameIndex = lambda: [self.rawMaterialRowIndex, 18]
        self.rawMaterialRateIndex = lambda: [self.rawMaterialRowIndex, 19]
        self.rawMaterialUnitPriceIndex = lambda: [self.rawMaterialRowIndex, 20]
        self.rawMaterialAnnualAmountIndex = lambda: [self.rawMaterialRowIndex, 21]
        self.rawMaterialTotalAmount = lambda: [2, 22]

        # Product Data
        self.productNamesIndex = lambda: [self.productDataRowIndex, 25]
        self.productFlowRateIndex = lambda: [self.productDataRowIndex, 26]
        self.productUnitPriceIndex = lambda: [self.productDataRowIndex, 27]
        self.productAnnualSaleIndex = lambda: [self.productDataRowIndex, 28]
        self.productTotalRevenueIndex =lambda: [2, 29]

    def equipmentDataCost(self, funcType, args):
        def addProcessIndex(arg):
            indices = siiSOTEA.processAreaIndex()

        def addProcessName(arg):
            self.equipEntryRowIndex += 1
            self.openpyxlsExcel.update_cell_value(f"N{self.openpyxl_equipEntryRow}", arg)
            indices = siiSOTEA.processAreaNamesIndex()

        def addProcessName_openpyxl(arg):
            self.equipEntryRowIndex += 1
            self.openpyxl_equipEntryRow += 1
            self.openpyxlsExcel.update_cell_value(f"N{self.openpyxl_equipEntryRow}", arg)

        def getProcessNames(arg):
            val = xw.Range(f"N4:N200").value
            val = [str(x) for x in val if str(x) not in ["nan", "None"]]
            return val
            
        def addInstalledCost(cost):
            self.openpyxlsExcel.update_cell_value(f"O{self.openpyxl_equipEntryRow}", cost)

        def addInstalledCost_openpyxl(arg):
            self.openpyxlsExcel.update_cell_value(f"O{self.openpyxl_equipEntryRow}", arg)
        
        def getInstalledCost(arg):
            val = xw.Range(f"O4:O200").value
            val = [float(x) for x in val if str(x) not in ["nan","None"]]
            return val

        def getSideBatteryLimit(arg):
            val = xw.Range(f"P4").value
            return val
        
        self.addProcessIndex = addProcessIndex
        self.addProcessName = addProcessName
        self.getProcessNames = getProcessNames
        self.addInstalledCost = addInstalledCost
        self.getInstalledCost = getInstalledCost
        self.getSideBatteryLimit = getSideBatteryLimit
        self.addProcessName_openpyxl = addProcessName_openpyxl
        self.addInstalledCost_openpyxl = addInstalledCost_openpyxl
        fn = getattr(self, funcType)
        return fn(args)

    def rawMaterialUtility(self, funcType, args):
        def addProcessIndex(arg):
            indices = siiSOTEA.processAreaIndex()

        def addRawMaterialName(arg):
            self.rawMaterialRowIndex += 1
            self.openpyxl_rawMaterialRow += 1
            self.openpyxlsExcel.update_cell_value(f"S{self.openpyxl_rawMaterialRow}", arg)

        def getRawMaterialName(arg):
            val = xw.Range(f"S4:S200").value
            val = [str(x) for x in val if str(x) not in ["nan", "None"]]
            return val
            
        def addRawMaterialRates(cost):
            self.openpyxlsExcel.update_cell_value(f"T{self.openpyxl_rawMaterialRow}", cost)
        
        def getRawMaterialRates(arg):
            val = xw.Range(f"T4:T200").value
            val = [format(float(x), 'f') for x in val if str(x) not in ["nan", "None"]]
            return val
        
        def addRawMaterialUnitPrice(cost):
            self.openpyxlsExcel.update_cell_value(f"U{self.openpyxl_rawMaterialRow}", cost)
        
        def getRawMaterialUnitPrice(arg):
            val = xw.Range(f"U4:U200").value
            val = [format(float(x), 'f') for x in val if str(x) not in ["nan", "None"]]
            return val
        
        def addRawMaterialAnnualAmount(arg): 
            temp = self.openpyxl_rawMaterialRow
            self.openpyxlsExcel.update_cell_value(f"V{temp}", f"=T{temp}*U{temp}*$B$10")

        def getRawMaterialAnnualAmount(arg):
            val = xw.Range(f"V{self.openpyxl_rawMaterialRow+1}:V200").value
            val = [format(float(x), 'f') for x in val if str(x) not in ["nan", "None"]]
            return val
        
        def getRawMaterialTotalAmount(arg):
            val = xw.Range(f"W4").value
            return format(float(val), 'f')
        
        def updateRawMaterialTotalAmount(arg): pass

        self.addProcessIndex = addProcessIndex
        self.addRawMaterialName = addRawMaterialName
        self.getRawMaterialName = getRawMaterialName
        self.addRawMaterialRates = addRawMaterialRates
        self.getRawMaterialRates = getRawMaterialRates
        self.addRawMaterialUnitPrice = addRawMaterialUnitPrice
        self.getRawMaterialUnitPrice = getRawMaterialUnitPrice
        self.addRawMaterialAnnualAmount = addRawMaterialAnnualAmount
        self.getRawMaterialAnnualAmount = getRawMaterialAnnualAmount
        self.getRawMaterialTotalAmount = getRawMaterialTotalAmount
        self.updateRawMaterialTotalAmount = updateRawMaterialTotalAmount
        fn = getattr(self, funcType)
        return fn(args)	

    def productsData(self, funcType, args):
        def addProductDataName(arg):
            self.productDataRowIndex += 1
            self.openpyxl_productDataRow += 1
            self.openpyxlsExcel.update_cell_value(f"Z{self.openpyxl_productDataRow}", arg)

        def getProductDataName(arg):
            val = xw.Range(f"Z4:Z200").value
            val = [str(x) for x in val if str(x) not in ["nan", "None"]]
            return val
            
        def addProductDataFlowRates(cost):
            self.openpyxlsExcel.update_cell_value(f"AA{self.openpyxl_productDataRow}", cost)
        
        def getProductDataRates(arg):
            val = xw.Range(f"AA4:AA200").value
            val = [format(float(x), 'f') for x in val if str(x) not in ["nan", "None"]]
            return val
        
        def addProductDataUnitPrice(cost):
            self.openpyxlsExcel.update_cell_value(f"AB{self.openpyxl_productDataRow}", cost)
        
        def getProductDataUnitPrice(arg):
            val = xw.Range(f"AB4:AB200").value
            val = [format(float(x), 'f') for x in val if str(x) not in ["nan", "None"]]
            return val

        def addProductDataAnnualSale(arg): 
            temp = self.openpyxl_productDataRow
            self.openpyxlsExcel.update_cell_value(f"AC{temp}", f"=AA{temp}*AB{temp}*B10")

        def getProductDataAnnualSale(arg):
            val = xw.Range(f"AC4:AC200").value
            val = [format(float(x), 'f') for x in val if str(x) != "None"]
            return val
        
        def getProductDataTotalRevenu(arg):
            val = xw.Range(f"AD4").value
            return format(float(val), 'f')		

        self.addProductDataName = addProductDataName
        self.getProductDataName = getProductDataName
        self.addProductDataFlowRates = addProductDataFlowRates
        self.getProductDataRates = getProductDataRates
        self.addProductDataUnitPrice = addProductDataUnitPrice
        self.getProductDataUnitPrice = getProductDataUnitPrice
        self.addProductDataAnnualSale = addProductDataAnnualSale
        self.getProductDataAnnualSale = getProductDataAnnualSale
        self.getProductDataTotalRevenu = getProductDataTotalRevenu
        fn = getattr(self, funcType)	
        return fn(args)		


# Example usage
CURRENT_FOLDER = os.path.dirname(__file__)
excel_file = os.path.join(CURRENT_FOLDER, "bin", "general_sheet.xlsx")
sheet_name = "100% Wind - 0% Solar"

##########################################################################################
#####################################  Writing Excel Data ################################
##########################################################################################

##################################################################
# Equipment Data -- Write Examples
excel = ExcelManipulator(excel_file, sheet_name)
siiSOTEA = SIISOTEA(df=None, excel=excel)
siiSOTEA.equipmentDataCost("addProcessName_openpyxl", "Test Equipment Data 1")
siiSOTEA.equipmentDataCost("addInstalledCost_openpyxl", float(1.0))

siiSOTEA.equipmentDataCost("addProcessName_openpyxl", "Test Equipment Data 2")
siiSOTEA.equipmentDataCost("addInstalledCost_openpyxl", float(2.0))

siiSOTEA.equipmentDataCost("addProcessName_openpyxl", "Test Equipment Data 3")
siiSOTEA.equipmentDataCost("addInstalledCost_openpyxl", float(3.0))

##################################################################
# Raw Material Data -- Write Examples
siiSOTEA.rawMaterialUtility("addRawMaterialName", "Testing Raw Material 1")
siiSOTEA.rawMaterialUtility("addRawMaterialRates", float(1.0))
siiSOTEA.rawMaterialUtility("addRawMaterialUnitPrice", float(0.1))
siiSOTEA.rawMaterialUtility("addRawMaterialAnnualAmount", "")


siiSOTEA.rawMaterialUtility("addRawMaterialName", "Testing Raw Material 2")
siiSOTEA.rawMaterialUtility("addRawMaterialRates", float(2.0))
siiSOTEA.rawMaterialUtility("addRawMaterialUnitPrice", float(0.2))
siiSOTEA.rawMaterialUtility("addRawMaterialAnnualAmount", "")

siiSOTEA.rawMaterialUtility("addRawMaterialName", "Testing Raw Material 3")
siiSOTEA.rawMaterialUtility("addRawMaterialRates", float(3.0))
siiSOTEA.rawMaterialUtility("addRawMaterialUnitPrice", float(0.3))
siiSOTEA.rawMaterialUtility("addRawMaterialAnnualAmount", "")

##################################################################
# Product Data -- Write Examples
siiSOTEA.productsData("addProductDataName", "Testing Product Data 1")
siiSOTEA.productsData("addProductDataFlowRates", float(1.0))
siiSOTEA.productsData("addProductDataUnitPrice", float(0.1))
siiSOTEA.productsData("addProductDataAnnualSale", "")

siiSOTEA.productsData("addProductDataName", "Testing Product Data 2")
siiSOTEA.productsData("addProductDataFlowRates", float(2.0))
siiSOTEA.productsData("addProductDataUnitPrice", float(0.2))
siiSOTEA.productsData("addProductDataAnnualSale", "")

siiSOTEA.productsData("addProductDataName", "Testing Product Data 3")
siiSOTEA.productsData("addProductDataFlowRates", float(3.0))
siiSOTEA.productsData("addProductDataUnitPrice", float(0.3))
siiSOTEA.productsData("addProductDataAnnualSale", "")

excel.save()
del siiSOTEA



##########################################################################################
#####################################  Reading Excel Data ################################
##########################################################################################
import xlwings as xw
wb = xw.Book(excel_file)
siiSOTEA = SIISOTEA(df=None, excel=None)

##################################################################
# Equipment Data Cost Examples
process = siiSOTEA.equipmentDataCost("getProcessNames", "")
cost = siiSOTEA.equipmentDataCost("getInstalledCost", "")
battery = siiSOTEA.equipmentDataCost("getSideBatteryLimit", "")
print(process)
print(cost)
print(battery)
print("\n")

###################################################################
# Raw Material / Utility Examples
RM_Names = siiSOTEA.rawMaterialUtility("getRawMaterialName", "")
RM_Rates = siiSOTEA.rawMaterialUtility("getRawMaterialRates", "")
RM_UnitPrice = siiSOTEA.rawMaterialUtility("getRawMaterialUnitPrice", "")
RM_AnnualAmount = siiSOTEA.rawMaterialUtility("getRawMaterialAnnualAmount", "")
RM_Tota = siiSOTEA.rawMaterialUtility("getRawMaterialTotalAmount", "")
print(RM_Names)
print(RM_Rates)
print(RM_UnitPrice)
print(RM_AnnualAmount)
print(RM_Tota)
print("\n")

###################################################################
# Product Data Examples
PD_Names = siiSOTEA.productsData("getProductDataName", "")
PD_Rates = siiSOTEA.productsData("getProductDataRates", "")
PD_UnitPrice = siiSOTEA.productsData("getProductDataUnitPrice", "")
PD_AnnualAmount = siiSOTEA.productsData("getProductDataAnnualSale", "")
PD_Total = siiSOTEA.productsData("getProductDataTotalRevenu", "")
print(PD_Names)
print(PD_Rates)
print(PD_UnitPrice)
print(PD_AnnualAmount)
print(PD_Total)
print("\n")

wb.close()
sys.exit()
