import os
import pandas as pd
from openpyxl import load_workbook

class ExcelManipulator:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.workbook = load_workbook(filename=file_path)
        self.sheet_name = sheet_name
        self.sheet = self.workbook.active if sheet_name is None else self.workbook[sheet_name]

    def read_cell(self, cell):
        return self.sheet[cell].value

    def write_cell(self, cell, value):
        self.sheet[cell].value = value

    def save(self):
        self.workbook.save(filename=self.file_path)

    def add_value_to_input_column(self, input_column, value):
        for row in range(1, self.sheet.max_row + 1):
            cell = self.sheet[f"{input_column}{row}"]
            if cell.value is not None: 
                continue
            cell.value = value
            return

    def get_value_from_output_cell(self, output_cell):
        return self.sheet[output_cell].value

    def vlookup(self, lookup_value, table_range, result_column):
        for row in self.sheet.iter_rows(min_row=self.sheet[table_range.split(':')[0]].row,
                max_row=self.sheet[table_range.split(':')[1]].row,
                min_col=self.sheet[table_range.split(':')[0]].column,
                max_col=self.sheet[table_range.split(':')[1]].column):
            if str(row[0].value) == lookup_value:
                return row[result_column - 1].value
        excel.save()
        return None
    
    def update_cell_value(self, cell, value):
        self.sheet[cell].value = value
        self.workbook.save(filename=self.file_path)


class PandasREAD:
    def __init__(self, path) -> None:
        self.workbook_path = path
        self.workbook = load_workbook(filename=path)
        self.sheet = self.workbook.active

    def getValueUsing_Formula(self, formula):
        self.sheet["J48"].value = formula               # "=SUM(A1:A10)" 
        self.workbook.save(filename=self.workbook_path)
        return self.readValue("J48")
    
    def getCurrent_FutureCost(self, lookupValue):
        import xlwings as xw

        app = xw.App(visible=False)
        wb = app.books.open(self.workbook_path)
        sheet = wb.sheets.active

        table_range = sheet.range('I5:L11').value
        df = pd.DataFrame(table_range, columns=['Year', 'Wind Turbine', 'Solar PV', 'Electrolyzer'])

        data_dict = {}
        for index, row in df.iterrows():
            year = str(int(row['Year']))
            data_dict[year] = {
                'Wind Turbine': str(row['Wind Turbine']),
                'Solar PV': str(row['Solar PV']),
                'Electrolyzer': str(row['Electrolyzer'])
            }
        result = None
        if lookupValue in data_dict.keys(): result = data_dict[lookup_value]
        wb.close()
        app.quit()
        return result

# Example usage
CURRENT_FOLDER = os.path.dirname(__file__)
excel_file = os.path.join(CURRENT_FOLDER, "bin", "InputData_Sheet.xlsx")
sheet_name = "Input Data"

lookup_value = input("Enter Year: ")
inputApproach = input("Input Approach (Lowest, Mid, Highest): ")

excel = ExcelManipulator(excel_file, sheet_name)
excel.update_cell_value("G15", inputApproach)
excel.save()

data = PandasREAD(excel_file).getCurrent_FutureCost(lookup_value)
print(data)


